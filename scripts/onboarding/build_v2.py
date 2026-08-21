#!/usr/bin/env python3
"""
Build the v2 budget workbook, designed around the Roster.

The v1 structure grew organically and Incomes ended up doing three jobs at
once: the payment ledger in A-F, a per-student roster block in M-R, and helper
name lists elsewhere. Dues Status then read that roster block **by row
position** — Incomes!M2, Incomes!S2 — so inserting a row in the wrong place
silently reattributes someone's money.

v2 separates them:

    Roster    the team. One row per student, participant id as the key,
              contacts, start date, and whether they are registered with FIRST.
              Source of truth for who exists.
    Incomes   the payment ledger, append-only, each row carrying the
              participant id it belongs to.
    Dues      computed only. Looks up students from Roster and sums payments
              from Incomes by participant id. Holds no data of its own.

Being on the Roster does not require FIRST registration — a student can be
paying dues before they register, and a student can register before their
first payment. FIRST status is a column, not a gate.

Nothing here touches the live workbook. It writes a new file for comparison.

Usage:
    python3 scripts/onboarding/build_v2.py <source.xlsx> --roster-csv <roster.csv> --out <v2.xlsx>
"""

import argparse
import csv
import json
import re
import sys
import unicodedata
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

LEDGER = Path.home() / "STEMC-onboarding" / "state.json"

# Per-student facts that no system can derive, loaded from outside the repo.
#
# Nicknames, mid-season start dates, and departure dates all name real minors,
# and this repository is public — so they live beside the onboarding ledger in
# ~/STEMC-onboarding/ rather than in this file. The script ships knowing the
# shape of that data, never the data itself.
#
#   aliases         what the budget calls someone -> their legal name. The
#                   one-time join that lets a nickname in the ledger meet a
#                   legal name in FIRST. Declared, never inferred: matching on
#                   last-name-plus-initial would also merge two real siblings,
#                   and declared beats inferred when the thing being merged is
#                   money.
#   startOverrides  students who joined mid-season and are billed from the
#                   month they started, not the season opening.
#   endDates        students who have left, and the last month they are billed
#                   for. They stay on the Roster: their payments are still in
#                   the ledger and still have to reconcile, and a season total
#                   that quietly changes when someone leaves cannot be audited.
OVERRIDES_PATH = Path.home() / "STEMC-onboarding" / "roster-overrides.json"


def load_overrides():
    if not OVERRIDES_PATH.exists():
        return {}, {}, {}
    raw = json.loads(OVERRIDES_PATH.read_text())
    d = lambda m: {k: date.fromisoformat(v) for k, v in (raw.get(m) or {}).items()}
    return (raw.get("aliases") or {}), d("startOverrides"), d("endDates")


ALIASES, START_OVERRIDES, END_DATES = load_overrides()

# Zeffy, read straight from the API rather than through the v1 importer.
#
# The v1 importer writes with openpyxl, which rewrites the whole workbook and
# drops the chart in 'FRC Current Budget'. It also resolves students against the
# roster block in Incomes column M, so a student who pays in full before being
# added to that block never imports at all. Reading Zeffy here instead leaves
# v1 and its chart alone, and matches against the Roster.
ZEFFY_BASE = "https://api.zeffy.com/api/v1"
ZEFFY_ENV = Path("/Volumes/Development/tiedyesamurai.org/.env")
SEASON_START = datetime(2026, 6, 1)

# Both teams take money through the same Zeffy account, so the campaign is the
# only thing separating Samurai income from Jedi income. Two of these were
# labelled 'MembershipV2 (unused)' in the older importer; they are not unused,
# they are the Jedi's. Titles below are Zeffy's own, from /campaigns.
#
# 'legacy' marks the 2025 form: money still arriving on it is arrears for last
# season, so it belongs in the ledger but not against this season's dues.
CAMPAIGNS = {
    "525d98ea-bd95-4c97-9290-e527e390ad9a": {
        "team": "samurai", "type": "Fee - Zeffy", "title": "Tie Dye Samurai Season Fees"},
    "1187f893-a6c8-41d4-be5d-6363e0606605": {
        "team": "samurai", "type": "Fee - legacy form", "legacy": True,
        "title": "Tie Dye Samurai 2025 FRC Season"},
    "c351057d-bee1-4f69-948f-601c50b1f932": {
        "team": "samurai", "type": "Donation", "title": "Tie Dye Samurai Needs Your Help"},
    "628d5252-792b-491f-ab69-26479d8e92d6": {
        "team": "jedi", "type": "Fee - Zeffy", "title": "Tie Dye Jedi Team Participation & Dues"},
    "686086a7-b581-48e6-9534-78c9b4771477": {
        "team": "jedi", "type": "Donation",
        "title": "Support the Tie Dye Jedi FIRST Robotics Team"},
}

# Ten payments on the 1st, July through April.
FIRST_PAYMENT = date(2026, 7, 1)
LAST_PAYMENT = date(2027, 4, 1)
MONTHLY = 215

HEADER_FILL = PatternFill("solid", fgColor="081F3F")
HEADER_FONT = Font(color="FFFFFF", bold=True)
NOTE_FONT = Font(italic=True, color="666666", size=9)


def payment_dates():
    out, y, m = [], FIRST_PAYMENT.year, FIRST_PAYMENT.month
    while (y, m) <= (LAST_PAYMENT.year, LAST_PAYMENT.month):
        out.append(date(y, m, 1))
        m += 1
        if m > 12:
            m, y = 1, y + 1
    return out


def zeffy_key():
    for line in ZEFFY_ENV.read_text().splitlines():
        if line.startswith("ZEFFY_API_KEY"):
            return line.split("=", 1)[1].strip().strip("\"'")
    sys.exit(f"\nZEFFY_API_KEY not found in {ZEFFY_ENV}\n")


def norm(s):
    """Loosen a free-text name enough to match. 'Alex Rivera- July' -> 'alex rivera'."""
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = s.split("-")[0]
    s = re.sub(r"\(.*?\)", " ", s)
    s = re.sub(r"[^a-zA-Z ]", " ", s)
    return re.sub(r"\s+", " ", s).strip().lower()


def fetch_zeffy(team):
    """
    Every payment on the fee and donation campaigns since the season opened.

    Returns (payments, subscriptions) where subscriptions maps a normalized
    student name to the most recent recurring object seen for them — which is
    where autopay comes from. Zeffy already knows who is on a subscription and
    whether it is still alive; asking it beats maintaining a column by hand.
    """
    import requests  # imported here so the script still runs without network deps

    s = requests.Session()
    s.headers["Authorization"] = f"Bearer {zeffy_key()}"
    since = SEASON_START.timestamp()

    # The API accepts campaign_id and then ignores it — passing one returns the
    # full set regardless. Filtering per campaign server-side therefore fetches
    # the same payments once per campaign and silently multiplies the ledger.
    # So: fetch once, classify here.
    payments, subs, skipped = [], {}, {}
    cursor = None
    while True:
        params = {"limit": 100}
        if cursor:
            params["starting_after"] = cursor
        j = s.get(f"{ZEFFY_BASE}/payments", params=params, timeout=30).json()
        for p in j["data"]:
            cfg = CAMPAIGNS.get(p.get("campaign_id"))
            if p["created"] < since or p.get("status") != "succeeded":
                continue
            # Other team's money is not this budget's income.
            if not cfg or cfg["team"] != team:
                if cfg:
                    skipped[cfg["title"]] = skipped.get(cfg["title"], 0) + p["amount"] / 100
                continue
            p["_campaign"] = cfg
            payments.append(p)
        if not j.get("has_more"):
            break
        cursor = j.get("next_cursor")

    for p in payments:
        rec = p.get("recurring")
        if not isinstance(rec, dict) or not rec.get("is_recurring"):
            continue
        who = norm(student_answer(p))
        if not who:
            continue
        prev = subs.get(who)
        # Keep the latest, but never let a later 'active' hide an earlier
        # cancellation on a different subscription — cancelled always wins,
        # because that is the one that needs attention.
        if prev and prev["status"] == "cancelled":
            continue
        if not prev or p["created"] > prev["_created"]:
            subs[who] = {**rec, "_created": p["created"]}

    return payments, subs, skipped


def student_answer(p):
    """The student a payment names, from the form's own question."""
    for item in p.get("items") or []:
        for q in item.get("questions") or []:
            if "student" in str(q.get("question", "")).lower():
                a = (q.get("answer") or "").strip()
                if a:
                    return a
    return ""


def read_zeffy_map(wb):
    """buyer email -> [(student, share), ...] for parents who pay under their own name."""
    if "Zeffy Map" not in wb.sheetnames:
        return {}
    out = defaultdict(list)
    ws = wb["Zeffy Map"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        email = (row[0] or "").strip().lower() if row[0] else ""
        student = (row[2] or "").strip() if len(row) > 2 and row[2] else ""
        if not email or not student:
            continue
        try:
            share = float(row[3]) if len(row) > 3 and row[3] is not None else 1.0
        except (TypeError, ValueError):
            share = 1.0
        out[email].append((student, share))
    return out


def key_of(full_name):
    """'Alexander Rivera' -> 'Alexander R'. The informal key v1 used."""
    parts = (full_name or "").strip().split()
    return f"{parts[0]} {parts[-1][:1]}" if len(parts) >= 2 else (full_name or "").strip()


def load_ledger_ids():
    """participant id keyed by full legal name, from the onboarding ledger."""
    if not LEDGER.exists():
        return {}
    data = json.loads(LEDGER.read_text()).get("participants", {})
    return {p["participantId"]: p for p in data.values() if p.get("participantId")}


def read_roster_csv(path):
    if not path:
        return {}
    with open(path, encoding="utf-8") as fh:
        return {r["Student"].strip(): r for r in csv.DictReader(fh) if r.get("Student")}


def read_roster_sheet(wb):
    """
    Fall back to the Roster sheet already pasted into the workbook.

    Saves re-exporting from FIRST just to rebuild, and means the contacts come
    from whatever was last reviewed rather than a stale CSV on disk.
    """
    if "Roster" not in wb.sheetnames:
        return {}
    ws = wb["Roster"]
    head = [str(c.value).strip() if c.value else "" for c in ws[1]]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(head, (("" if v is None else str(v).strip()) for v in row)))
        if rec.get("Student"):
            out[rec["Student"]] = rec
    return out


def collect_students(src, roster_csv):
    """
    Everyone who is on the team, from either source.

    The budget knows who is paying; FIRST knows who has registered. Neither is
    complete on its own — v1 had eight students billed but not registered, and
    two registered but not billed.
    """
    wb = openpyxl.load_workbook(src, data_only=True)
    inc = wb["Incomes"]

    if not roster_csv:
        roster_csv = read_roster_sheet(wb)

    students = {}

    # From the v1 roster block inside Incomes.
    for r in range(2, 40):
        name = inc.cell(r, 13).value
        if not name or str(name).strip() in ("0", ""):
            continue
        name = str(name).strip()
        # Resolve the nickname now, so the FIRST record below lands on the same
        # student instead of creating a second row for the same kid.
        name = ALIASES.get(name.lower(), name)
        students[name] = {"name": name, "in_budget": True, "in_first": False}

    # From the FIRST/Jotform roster export.
    for name, row in roster_csv.items():
        s = students.setdefault(name, {"name": name, "in_budget": False})
        s["in_first"] = True
        s.update(
            {
                "participant_id": row.get("Participant ID", ""),
                "grade": row.get("Grade", ""),
                "parent1": row.get("Parent 1", ""),
                "parent1_phone": row.get("Parent 1 Phone", ""),
                "parent1_email": row.get("Parent 1 Email", ""),
                "parent2": row.get("Parent 2", ""),
                "parent2_phone": row.get("Parent 2 Phone", ""),
                "parent2_email": row.get("Parent 2 Email", ""),
                "student_email": row.get("Student Email", ""),
                "student_phone": row.get("Student Phone", ""),
                "packet": row.get("Packet", ""),
                # Carried through on a rebuild so hand-entered values survive.
                "autopay": row.get("Autopay", ""),
            }
        )

    ledger = load_ledger_ids()

    # Anyone without an id gets the next one. Being billed but not yet
    # registered is a normal state, not an error.
    used = {
        int(m.group(1))
        for s in students.values()
        if (m := re.search(r"-(\d+)$", s.get("participant_id") or ""))
    }
    nxt = max(used, default=0)
    for name in sorted(students):
        if not students[name].get("participant_id"):
            nxt += 1
            students[name]["participant_id"] = f"TDS-27-{nxt:03d}"
        # Start date, most authoritative first: the onboarding ledger (set with
        # `onboard.mjs --set-start`, so adding a student never means editing
        # this script), then the table above for people who predate the ledger
        # carrying dates, then the season opening for everyone carried over.
        pid = students[name]["participant_id"]
        ledger_start = (ledger.get(pid) or {}).get("startDate")
        if ledger_start:
            students[name]["start"] = date.fromisoformat(ledger_start)
        else:
            students[name]["start"] = START_OVERRIDES.get(name, FIRST_PAYMENT)

    return wb, sorted(students.values(), key=lambda s: s["participant_id"])


ROSTER_COLS = [
    ("Participant ID", 14),
    ("Student", 22),
    ("Key", 14),
    ("Grade", 7),
    ("Start Date", 12),
    ("End Date", 12),
    ("Autopay", 10),
    ("FIRST", 10),
    ("Packet", 11),
    ("Parent 1", 20),
    ("Parent 1 Phone", 15),
    ("Parent 1 Email", 28),
    ("Parent 2", 20),
    ("Parent 2 Phone", 15),
    ("Parent 2 Email", 28),
    ("Student Email", 26),
    ("Student Phone", 15),
    ("Notes", 24),
]

INCOME_COLS = [
    ("Date", 12),
    ("Participant ID", 14),
    ("Payer", 20),
    ("Type", 16),
    ("Amount", 10),
    ("Zeffy Payment ID", 38),
    ("Notes", 24),
]

DUES_COLS = [
    ("Participant ID", 14),
    ("Student", 22),
    ("Start Date", 12),
    ("End Date", 12),
    ("Payments Owed", 14),
    ("Owed to Date", 13),
    ("Paid to Date", 13),
    ("Over / Under", 13),
    ("Months Behind", 14),
    ("Last Payment", 13),
    ("Days Since", 11),
    ("Autopay", 9),
    ("Status", 12),
    ("Action", 22),
]


def style_header(ws, cols):
    for i, (name, width) in enumerate(cols, start=1):
        c = ws.cell(row=1, column=i, value=name)
        c.fill, c.font = HEADER_FILL, HEADER_FONT
        c.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("source", help="v1 workbook to migrate from")
    ap.add_argument("--roster-csv", help="roster_sheet.py output, for contacts")
    ap.add_argument("--out", required=True)
    ap.add_argument("--zeffy", action="store_true",
                    help="pull payments and autopay live from Zeffy instead of copying v1's ledger")
    ap.add_argument("--team", default="samurai", choices=("samurai", "jedi"),
                    help="whose budget this is; the other team's Zeffy income is excluded")
    args = ap.parse_args()

    roster_csv = read_roster_csv(args.roster_csv)
    src, students = collect_students(args.source, roster_csv)
    # Payments in v1 are tagged with the informal key — 'Alex R'. Index the
    # aliases too, or a nickname's payments would not follow them to the row
    # filed under their legal name.
    by_key = {key_of(s["name"]): s for s in students}
    for alias, legal in ALIASES.items():
        target = next((s for s in students if s["name"] == legal), None)
        if target:
            by_key[key_of(alias.title())] = target

    # Full names, loosened, for matching what people type on the Zeffy form.
    by_name = {norm(s["name"]): s for s in students}
    for alias, legal in ALIASES.items():
        target = next((s for s in students if s["name"] == legal), None)
        if target:
            by_name[norm(alias)] = target

    zeffy_payments, subscriptions, other_team = ([], {}, {})
    if args.zeffy:
        zeffy_payments, subscriptions, other_team = fetch_zeffy(args.team)
        for who, rec in subscriptions.items():
            s = by_name.get(who)
            if s:
                s["autopay"] = rec["status"]        # active / cancelled
                s["_sub_interval"] = rec.get("interval", "")

    out = openpyxl.Workbook()
    out.remove(out.active)

    # ── Roster ──────────────────────────────────────────────────────────────
    ws = out.create_sheet("Roster")
    style_header(ws, ROSTER_COLS)
    for r, s in enumerate(students, start=2):
        start = s.get("start")
        if isinstance(start, datetime):
            start = start.date()
        ws.cell(r, 1, s["participant_id"])
        ws.cell(r, 2, s["name"])
        ws.cell(r, 3, key_of(s["name"]))
        ws.cell(r, 4, s.get("grade") or None)
        ws.cell(r, 5, start or FIRST_PAYMENT).number_format = "yyyy-mm-dd"
        end = END_DATES.get(s["name"])
        if end:
            ws.cell(r, 6, end).number_format = "yyyy-mm-dd"
        # Set from Zeffy's subscription status on every rebuild, so there is no
        # hand-maintained fallback to disagree with it.
        ws.cell(r, 7, s.get("autopay") or "")
        ws.cell(r, 8, "registered" if s.get("in_first") else "not yet")
        ws.cell(r, 9, s.get("packet") or ("—" if not s.get("in_first") else ""))
        ws.cell(r, 10, s.get("parent1") or "")
        ws.cell(r, 11, s.get("parent1_phone") or "")
        ws.cell(r, 12, s.get("parent1_email") or "")
        ws.cell(r, 13, s.get("parent2") or "")
        ws.cell(r, 14, s.get("parent2_phone") or "")
        ws.cell(r, 15, s.get("parent2_email") or "")
        ws.cell(r, 16, s.get("student_email") or "")
        ws.cell(r, 17, s.get("student_phone") or "")
        notes = []
        if not s.get("in_budget"):
            notes.append("new — not yet in v1 dues")
        if end:
            notes.append(f"left the team, billed through {end:%b %Y}")
        ws.cell(r, 18, "; ".join(notes))

    # Autopay comes from Zeffy's subscription status, so the values are Zeffy's
    # own: active, cancelled, or blank for no subscription at all. Constrained
    # to that list because the Dues formulas test for exactly these strings.
    dv = DataValidation(type="list", formula1='"active,cancelled"', allow_blank=True)
    dv.prompt = "Set from Zeffy on every rebuild — edits here are overwritten."
    dv.promptTitle = "Autopay"
    ws.add_data_validation(dv)
    dv.add(f"G2:G{len(students) + 1}")

    # ── Incomes ─────────────────────────────────────────────────────────────
    ws = out.create_sheet("Incomes")
    style_header(ws, INCOME_COLS)
    row = 2
    unmatched = []

    if args.zeffy:
        zmap = read_zeffy_map(src)
        for p in sorted(zeffy_payments, key=lambda p: p["created"]):
            when = datetime.fromtimestamp(p["created"]).date()
            amount = p["amount"] / 100
            buyer = p.get("buyer") or {}
            email = (buyer.get("email") or "").strip().lower()
            typed = student_answer(p)

            # Who this is for: what the form was told, else a Zeffy Map rule for
            # this buyer. A parent paying for two children splits by share, so
            # one transaction becomes one row per child.
            if typed:
                targets = [(typed, 1.0)]
            elif email in zmap:
                targets = zmap[email]
            else:
                targets = [("", 1.0)]

            legacy = p["_campaign"].get("legacy", False)
            for name, share in targets:
                match = by_name.get(norm(name)) if name else None
                if name and not match:
                    unmatched.append((when, name, amount * share))
                note = ""
                if legacy:
                    note = "2025-26 arrears, not this season"
                elif not name:
                    who_hint = f"{buyer.get('first_name','?')[:1]}.{buyer.get('last_name','?')[:1]}."
                    note = f"no student named — needs a Zeffy Map rule for {who_hint}"
                elif not match:
                    note = f"no roster match for {name!r}"
                elif share != 1.0:
                    note = f"{share:.0%} of a combined payment"

                ws.cell(row, 1, when).number_format = "yyyy-mm-dd"
                # Arrears and unattributable money get no id on purpose: the
                # Dues SUMIFS keys on the id, so a blank keeps them out of this
                # season's balances while still showing up in the ledger total.
                ws.cell(row, 2, match["participant_id"] if (match and not legacy) else "")
                ws.cell(row, 3, name or buyer.get("full_name") or "(unidentified)")
                ws.cell(row, 4, p["_campaign"]["type"])
                ws.cell(row, 5, round(amount * share, 2)).number_format = '"$"#,##0.00'
                ws.cell(row, 6, p["id"])
                ws.cell(row, 7, note)
                row += 1
    else:
        inc = src["Incomes"]
        for r in range(2, inc.max_row + 1):
            d, who, typ, amt = (inc.cell(r, c).value for c in (1, 2, 3, 4))
            if not d or amt in (None, "", 0):
                continue
            note, pid = inc.cell(r, 5).value, inc.cell(r, 6).value
            match = by_key.get(str(who).strip()) if who else None
            if who and not match:
                unmatched.append((d, str(who), amt))
            ws.cell(row, 1, d.date() if isinstance(d, datetime) else d).number_format = "yyyy-mm-dd"
            ws.cell(row, 2, match["participant_id"] if match else "")
            ws.cell(row, 3, str(who) if who else "")
            ws.cell(row, 4, typ or "")
            ws.cell(row, 5, amt).number_format = '"$"#,##0.00'
            ws.cell(row, 6, pid or "")
            ws.cell(row, 7, note or "")
            row += 1

    # ── Dues ────────────────────────────────────────────────────────────────
    ws = out.create_sheet("Dues")
    style_header(ws, DUES_COLS)
    sched = payment_dates()
    last = len(students) + 1
    N = len(sched) + 1
    # A student who has left is billed to their end date; everyone else to the
    # end of the season. A far-future date stands in for "still here" so one
    # formula covers both without a special case per row.
    cap = lambda r: f'IF($D{r}="",DATE(2099,1,1),$D{r})'
    window = lambda r: (
        f'--(Schedule!$A$2:$A${N}>=EOMONTH($C{r},-1)+1),'
        f'--(Schedule!$A$2:$A${N}<={cap(r)})'
    )
    for r, s in enumerate(students, start=2):
        ws.cell(r, 1, f"=Roster!A{r}")
        ws.cell(r, 2, f"=Roster!B{r}")
        ws.cell(r, 3, f"=Roster!E{r}").number_format = "yyyy-mm-dd"
        ws.cell(r, 4, f"=Roster!F{r}").number_format = "yyyy-mm-dd"
        # Payments they owe for the whole time they are on the team.
        ws.cell(r, 5, f'=SUMPRODUCT({window(r)})')
        # Of those, the ones whose due date has passed.
        ws.cell(
            r, 6,
            f'=SUMPRODUCT({window(r)},--(Schedule!$A$2:$A${N}<=TODAY()))*{MONTHLY}',
        ).number_format = '"$"#,##0'
        ws.cell(r, 7, f'=SUMIFS(Incomes!$E:$E,Incomes!$B:$B,$A{r})').number_format = '"$"#,##0'
        ws.cell(r, 8, f"=G{r}-F{r}").number_format = '"$"#,##0;[Red]-"$"#,##0'
        ws.cell(r, 9, f'=IF(H{r}<0,ROUND(-H{r}/{MONTHLY},1),0)')
        # MAXIFS post-dates the file format, so a formula written straight into
        # the XML has to carry the _xlfn. prefix or Excel reads it as #NAME?.
        # Excel strips the prefix on display; v1 carries it for the same reason.
        ws.cell(r, 10, f'=IFERROR(IF(_xlfn.MAXIFS(Incomes!$A:$A,Incomes!$B:$B,$A{r})=0,"",'
                       f'_xlfn.MAXIFS(Incomes!$A:$A,Incomes!$B:$B,$A{r})),"")').number_format = "yyyy-mm-dd"
        ws.cell(r, 11, f'=IF(J{r}="","",TODAY()-J{r})')
        ws.cell(r, 12, f"=Roster!G{r}")
        ws.cell(
            r, 13,
            f'=IF($D{r}<>"","Left",'
            f'IF(H{r}>=0,"Current",IF(I{r}<=1,"Watch",IF(I{r}<=2,"Behind","CRITICAL"))))',
        )
        # What to actually do. Order matters more than it looks:
        #
        # A student who has left is settled against what they owed up to their
        # end date, so the same overpayment that reads "paid ahead" while they
        # are on the team becomes a refund the moment they are not.
        #
        # Then cancelled autopay, tested before the balance — someone who
        # cancelled while paid ahead looks fine today and stops paying
        # tomorrow, and checking the balance first hides exactly that case.
        #
        # Only then a shortfall, which means different things depending on
        # whether a subscription is still running. Texting a payment reminder
        # to a family whose autopay is working is how you annoy the parents who
        # already did the right thing.
        ws.cell(
            r, 14,
            f'=IF($D{r}<>"",IF(H{r}>0,"left — refund "&TEXT(H{r},"$#,##0"),'
            f'IF(H{r}<0,"left — owes "&TEXT(-H{r},"$#,##0"),"left — settled")),'
            f'IF(AND(L{r}="cancelled",E{r}*{MONTHLY}>G{r}),"autopay cancelled — restart it",'
            f'IF(H{r}>=0,"—",'
            f'IF(L{r}="active","autopay active — due to catch up",'
            f'"no autopay — remind parent 1"))))',
        )
    for col in (6, 7, 8):
        ws.cell(last + 2, col, f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{last})")
        ws.cell(last + 2, col).number_format = '"$"#,##0;[Red]-"$"#,##0'
    ws.cell(last + 2, 2, "Totals").font = Font(bold=True)

    # ── Schedule ────────────────────────────────────────────────────────────
    # The payment calendar as data rather than buried in a formula, so changing
    # the season means editing a column, not rewriting every row.
    ws = out.create_sheet("Schedule")
    style_header(ws, [("Payment Date", 14), ("Amount", 10)])
    for r, d in enumerate(sched, start=2):
        ws.cell(r, 1, d).number_format = "yyyy-mm-dd"
        ws.cell(r, 2, MONTHLY).number_format = '"$"#,##0'
    ws.cell(len(sched) + 3, 1, "Season total").font = Font(bold=True)
    ws.cell(len(sched) + 3, 2, f"=SUM(B2:B{len(sched)+1})").number_format = '"$"#,##0'

    # ── Carried across unchanged ────────────────────────────────────────────
    # Only Zeffy Map for now — it is what feeds Incomes. Expenses and the rest
    # stay in v1 until we work out how their data gets in.
    for name in ("Zeffy Map",):
        if name not in src.sheetnames:
            continue
        s_ws, d_ws = src[name], out.create_sheet(name)
        for r in s_ws.iter_rows():
            for c in r:
                if c.value is not None:
                    d_ws.cell(c.row, c.column, c.value)

    Path(args.out).parent.mkdir(parents=True, exist_ok=True)
    out.save(args.out)

    billed_not_reg = [s["name"] for s in students if s["in_budget"] and not s.get("in_first")]
    reg_not_billed = [s["name"] for s in students if s.get("in_first") and not s["in_budget"]]

    print(f"\nWrote {args.out}")
    print(f"  Roster    {len(students)} students")
    print(f"  Incomes   {row - 2} payments" + (f", {len(unmatched)} unattributed" if unmatched else ""))
    print(f"  Dues      computed from Roster + Incomes by participant id")
    print(f"  Schedule  {len(sched)} payments of ${MONTHLY}")
    if billed_not_reg:
        print(f"\n  Billed but not registered with FIRST ({len(billed_not_reg)}):")
        print("   ", ", ".join(billed_not_reg))
    if reg_not_billed:
        print(f"\n  Registered but not yet in the dues ledger ({len(reg_not_billed)}):")
        print("   ", ", ".join(reg_not_billed))
    if unmatched:
        print(f"\n  Payments that could not be attributed to a student ({len(unmatched)}):")
        for when, who, amt in unmatched[:12]:
            print(f"    {when}  ${amt:>8,.2f}  {who!r}")
        print("    These are in the ledger with no id, so they are visible but")
        print("    counted against nobody. Add a rule to the Zeffy Map sheet.")

    if args.zeffy and other_team:
        print(f"\n  Excluded — belongs to the other team, not this budget:")
        for title, amt in sorted(other_team.items(), key=lambda kv: -kv[1]):
            print(f"    ${amt:>9,.2f}  {title}")

    if args.zeffy:
        auto = defaultdict(list)
        for s in students:
            auto[s.get("autopay") or "none"].append(s["name"])
        print("\n  Autopay, from Zeffy:")
        for status in ("active", "cancelled", "none"):
            if auto.get(status):
                print(f"    {status:<10} {len(auto[status]):>2}  {', '.join(sorted(auto[status]))}")


if __name__ == "__main__":
    main()

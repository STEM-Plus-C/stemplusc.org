#!/usr/bin/env python3
"""
Who owes what, computed from the data workbook.

One implementation, imported by everything that needs the answer. Before this
existed the sum lived in three places — the sync's report, the Dues query, and
the text drafter — and three copies of an arithmetic rule is three chances to
tell a family a different number than the sheet does.

The Excel side (powerquery/Dues.m) is unavoidably a fourth copy, because Excel
cannot import Python. It is kept deliberately identical, and this docstring is
the reminder to change both.

Reads only. Nothing here writes to a workbook.

    from dues import compute
    for d in compute():
        print(d["participant_id"], d["status"], d["balance"])
"""

from datetime import date
from pathlib import Path

import openpyxl

DATA_BOOK = Path.home() / "Library/CloudStorage/OneDrive-STEM+C/FRC Data 2026-2027.xlsx"

MONTHLY = 215
SEASON_END = date(2027, 4, 1)   # last scheduled payment


def _as_date(v):
    return v.date() if hasattr(v, "date") else v


def months_inclusive(a, b):
    """Payments due from month a to month b, counting both ends.

    A student who joins on the 1st owes for that month, so the start month
    counts. Off by one here is off by $215 per family.
    """
    return (b.year - a.year) * 12 + (b.month - a.month) + 1


def compute(book=None, today=None):
    """
    One record per student on the Roster.

    Payments with no participant id are excluded on purpose: they are last
    season's arrears and money nobody could attribute, and both belong in the
    ledger total without landing on a student's balance.
    """
    book = Path(book) if book else DATA_BOOK
    today = today or date.today()
    wb = openpyxl.load_workbook(book, data_only=True)

    paid, last_payment = {}, {}
    for row in wb["Incomes"].iter_rows(min_row=2, values_only=True):
        pid = str(row[1]).strip() if row[1] else ""
        if not pid or row[4] is None:
            continue
        paid[pid] = paid.get(pid, 0.0) + float(row[4])
        when = _as_date(row[0])
        if when and (pid not in last_payment or when > last_payment[pid]):
            last_payment[pid] = when

    head = [c.value for c in wb["Roster"][1]]
    out = []
    for row in wb["Roster"].iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        r = dict(zip(head, row))
        pid = str(r["Participant ID"]).strip()
        start = _as_date(r.get("Start Date"))
        end = _as_date(r.get("End Date"))

        # Billed from the month they started to the month they left, or to the
        # end of the season. A student who has left keeps their row: their
        # payments still have to reconcile, and a season total that quietly
        # changes when someone leaves cannot be audited.
        cap = min(end or date(2099, 1, 1), today, SEASON_END)
        owed = 0 if (not start or cap < start) else months_inclusive(start, cap) * MONTHLY
        p = paid.get(pid, 0.0)
        balance = p - owed

        if owed == 0:
            status = "Not yet billed"
        elif balance >= 0:
            status = "Current"
        elif -balance <= MONTHLY:
            status = "Watch"
        elif -balance <= MONTHLY * 2:
            status = "Behind"
        else:
            status = "CRITICAL"

        out.append(
            {
                "participant_id": pid,
                "full_name": (r.get("Student") or "").strip(),
                "key": (r.get("Key") or "").strip(),
                "start": start,
                "end": end,
                "owed": owed,
                "paid": p,
                "balance": balance,
                "months_behind": round(-balance / MONTHLY, 1) if balance < 0 else 0,
                "last_payment": last_payment.get(pid),
                "status": status,
                # Contacts travel with the record so callers do not have to
                # re-read a FIRST export just to find a phone number.
                "parent": (r.get("Parent 1") or "").strip(),
                "parent_phone": (r.get("Parent 1 Phone") or "").strip(),
                "parent2": (r.get("Parent 2") or "").strip(),
                "parent2_phone": (r.get("Parent 2 Phone") or "").strip(),
                "student_phone": (r.get("Student Phone") or "").strip(),
                "in_first": str(r.get("FIRST") or "").strip().lower() in ("yes", "true", "1"),
            }
        )

    return sorted(out, key=lambda d: d["participant_id"])


def season_total(record):
    """What this student owes for the whole season, not just to date."""
    if not record["start"]:
        return 0
    end = min(record["end"] or SEASON_END, SEASON_END)
    if end < record["start"]:
        return 0
    return months_inclusive(record["start"], end) * MONTHLY

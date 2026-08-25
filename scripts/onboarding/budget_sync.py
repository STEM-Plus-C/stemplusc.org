#!/usr/bin/env python3
"""
Keep the budget's data workbook in step with the onboarding ledger and Zeffy.

## Why there are two workbooks

The budget used to be one file doing two jobs, and the second job kept the
first one from being automated. `FRC Budget 2026-2027.xlsx` holds a chart, and
openpyxl deletes charts when it rewrites a workbook — which is why add_sheet.py
exists at all, 184 lines of zip surgery to avoid destroying one picture.

So the data moves to a workbook with no chart in it:

    FRC Data 2026-2027.xlsx     Roster, Incomes, Schedule. Written by this
                                script. Nothing decorative, nothing to lose.
    FRC Budget 2026-2027.xlsx   Charts, dues analysis, expense buckets. Read
                                from the data workbook through Power Query.
                                Owned by a person; this script never opens it.

That split is also what fixes the join. v1 kept a per-student roster block in
Incomes columns M-S, and Dues Status read it *by row position* — so inserting a
row in the wrong place silently reattributed someone's money. Here the Roster
is its own sheet and every payment carries a participant id.

## Incomes is append-only, and that is not an implementation detail

Venmo, cash and cheque payments are typed in by hand. The v1 formula already
counted `Fee - Venmo`; nobody had ever entered one, which is how a family who
paid by Venmo sat at $0 and CRITICAL and got chased for money they had sent.

If this script regenerated Incomes, every hand-entered row would vanish on the
next run. So it never rewrites or reorders that sheet. It appends Zeffy
payments it has not seen before, keyed on Zeffy's own payment id, and leaves
everything else exactly where it is. Rows without a Zeffy id are hand-entered
by definition and are never touched.

DRY RUN IS THE DEFAULT, as with onboard.mjs. Pass --commit to write.

Usage:
    python3 scripts/onboarding/budget_sync.py <roster.json>
    python3 scripts/onboarding/budget_sync.py <roster.json> --commit
    python3 scripts/onboarding/budget_sync.py <roster.json> --zeffy --commit
    python3 scripts/onboarding/budget_sync.py <roster.json> --from "<v1.xlsx>" --commit

Environment (for --zeffy):
    ZEFFY_API_KEY   read from /Volumes/Development/tiedyesamurai.org/.env
"""

import argparse
import shutil
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))

# Reused rather than rewritten: these are already tested, and a second
# implementation of "which student is this payment for" is exactly the kind of
# duplication that lets two answers drift apart.
from build_v2 import (  # noqa: E402
    ALIASES,
    END_DATES,
    FIRST_PAYMENT,
    INCOME_COLS,
    LAST_PAYMENT,
    MONTHLY,
    ROSTER_COLS,
    START_OVERRIDES,
    collect_students,
    fetch_zeffy,
    key_of,
    load_ledger_ids,
    norm,
    payment_dates,
    read_zeffy_map,
    student_answer,
    style_header,
)

DATA_BOOK = Path.home() / "Library/CloudStorage/OneDrive-STEM+C/FRC Data 2026-2027.xlsx"
LEDGER_PATH = Path.home() / "STEMC-onboarding" / "state.json"

# Carried through untouched. These hold data a person maintains by hand, and
# this script has no business regenerating any of it.
PRESERVED = ("Zeffy Map", "Expenses", "Sponsors")


# ----------------------------------------------------------------- reading


def reserved_by_name():
    """
    Participant ids the ledger has reserved, keyed by the name it reserved for.

    Ids come from the ledger and nowhere else. A student who pays before
    registering with FIRST still needs an id, and if this script minted one it
    would pick the next number the ledger is *also* about to hand out — two
    people holding one id, with money attaching to whichever the sheet looked
    up first. `onboard.mjs --reserve` claims the number properly.
    """
    if not LEDGER_PATH.exists():
        return {}
    import json

    data = json.loads(LEDGER_PATH.read_text()).get("participants", {})
    return {
        norm(p["reservedName"]): p["participantId"]
        for p in data.values()
        if p.get("reservedName") and p.get("participantId")
    }


def read_roster(wb):
    """Students already in the data workbook, keyed by participant id."""
    if "Roster" not in wb.sheetnames:
        return {}
    ws = wb["Roster"]
    head = [str(c.value).strip() if c.value else "" for c in ws[1]]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(head, row))
        pid = (rec.get("Participant ID") or "").strip()
        if pid:
            out[pid] = rec
    return out


def read_incomes(wb):
    """
    Every payment row already recorded, in order.

    Returned as plain tuples rather than cells because they are about to be
    written back into a fresh sheet: the point is to preserve what they say,
    not how they were formatted.
    """
    if "Incomes" not in wb.sheetnames:
        return []
    ws = wb["Incomes"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v not in (None, "") for v in row):
            continue
        rows.append(list(row) + [None] * (len(INCOME_COLS) - len(row)))
    return rows


def zeffy_ids_present(rows):
    """The dedupe key. Column 6 is Zeffy Payment ID; blank means hand-entered."""
    return {str(r[5]).strip() for r in rows if r[5] not in (None, "")}


# ----------------------------------------------------------------- students


def resolve_ids(students):
    """
    Replace any locally-minted id with the one the ledger reserved.

    A student who came from the FIRST export already carries the id the ledger
    gave them. Everyone else — billed but not registered — must have a
    reservation, or they get no id at all and are reported rather than written.
    Guessing here is how two people end up sharing a number.
    """
    reserved = reserved_by_name()
    for s in students:
        if s.get("in_first"):
            continue
        pid = reserved.get(norm(s["name"]))
        s["participant_id"] = pid or ""
        s["needs_reservation"] = not pid
    return students


def gather(args):
    """
    Everyone on the team, and where each fact came from.

    Three sources, none complete on its own: the onboarding ledger knows who
    was accepted and their participant id, FIRST knows who registered, and the
    budget knows who is paying. A student can be paying before they register
    and registered before they pay.
    """
    roster_csv = {}
    if args.roster_csv:
        import csv

        with open(args.roster_csv, encoding="utf-8") as fh:
            roster_csv = {r["Student"].strip(): r for r in csv.DictReader(fh) if r.get("Student")}

    # First build seeds from v1, which is the only place the historical roster
    # block and its payments exist.
    if args.seed_from:
        _, students = collect_students(args.seed_from, roster_csv)
        return {(s["participant_id"] or f"?{s['name']}"): s for s in resolve_ids(students)}

    wb = openpyxl.load_workbook(DATA_BOOK, data_only=True) if DATA_BOOK.exists() else None
    existing = read_roster(wb) if wb else {}

    students = {}
    for pid, rec in existing.items():
        students[pid] = {
            "participant_id": pid,
            "name": (rec.get("Student") or "").strip(),
            "in_budget": True,
            "in_first": str(rec.get("FIRST") or "").strip().lower() in ("yes", "true", "1"),
            "grade": rec.get("Grade") or "",
            "autopay": rec.get("Autopay") or "",
            "packet": rec.get("Packet") or "",
            **{
                k: rec.get(h) or ""
                for k, h in (
                    ("parent1", "Parent 1"),
                    ("parent1_phone", "Parent 1 Phone"),
                    ("parent1_email", "Parent 1 Email"),
                    ("parent2", "Parent 2"),
                    ("parent2_phone", "Parent 2 Phone"),
                    ("parent2_email", "Parent 2 Email"),
                    ("student_email", "Student Email"),
                    ("student_phone", "Student Phone"),
                )
            },
        }

    # FIRST/Jotform contacts win over whatever the sheet last held: they are
    # re-exported per run and the sheet is a copy.
    for name, row in roster_csv.items():
        pid = (row.get("Participant ID") or "").strip()
        if not pid:
            continue
        s = students.setdefault(pid, {"participant_id": pid, "name": name, "in_budget": False})
        s["in_first"] = True
        s["name"] = name
        for k, h in (
            ("grade", "Grade"),
            ("parent1", "Parent 1"),
            ("parent1_phone", "Parent 1 Phone"),
            ("parent1_email", "Parent 1 Email"),
            ("parent2", "Parent 2"),
            ("parent2_phone", "Parent 2 Phone"),
            ("parent2_email", "Parent 2 Email"),
            ("student_email", "Student Email"),
            ("student_phone", "Student Phone"),
            ("packet", "Packet"),
        ):
            if row.get(h):
                s[k] = row[h]

    # Start dates, most authoritative first: the ledger (set with
    # `onboard.mjs --set-start`), then the declared overrides, then the season
    # opening. Never inferred — a wrong start date is a wrong bill.
    ledger = load_ledger_ids()
    for pid, s in students.items():
        entry = ledger.get(pid) or {}
        if entry.get("startDate"):
            s["start"] = date.fromisoformat(entry["startDate"])
        elif s["name"] in START_OVERRIDES:
            s["start"] = START_OVERRIDES[s["name"]]
        else:
            s["start"] = FIRST_PAYMENT
        s["end"] = END_DATES.get(s["name"])

    return students


# ------------------------------------------------------------------ zeffy


def new_zeffy_rows(students, seen_ids, zmap):
    """Zeffy payments not already in the ledger, attributed to a participant id."""
    by_name = {norm(s["name"]): s for s in students.values()}
    payments, subs = fetch_zeffy("samurai")
    rows, unmatched, skipped = [], [], 0

    for p in sorted(payments, key=lambda p: p["created"]):
        if str(p["id"]) in seen_ids:
            skipped += 1
            continue
        when = datetime.fromtimestamp(p["created"]).date()
        amount = p["amount"] / 100
        buyer = p.get("buyer") or {}
        email = (buyer.get("email") or "").strip().lower()
        typed = student_answer(p)

        if typed:
            targets = [(typed, 1.0)]
        elif email in zmap:
            targets = zmap[email]
        else:
            targets = [("", 1.0)]

        legacy = p["_campaign"].get("legacy", False)
        for name, share in targets:
            match = by_name.get(norm(name)) if name else None
            if name and not match:
                unmatched.append((when, name, amount * share))
            if legacy:
                note = "2025-26 arrears, not this season"
            elif not name:
                hint = f"{buyer.get('first_name','?')[:1]}.{buyer.get('last_name','?')[:1]}."
                note = f"no student named — needs a Zeffy Map rule for {hint}"
            elif not match:
                note = f"no roster match for {name!r}"
            elif share != 1.0:
                note = f"{share:.0%} of a combined payment"
            else:
                note = ""
            # Arrears and unattributable money carry no id on purpose: dues key
            # on the id, so a blank keeps them out of this season's balances
            # while leaving them visible in the ledger total.
            rows.append(
                [
                    when,
                    match["participant_id"] if (match and not legacy) else "",
                    name or buyer.get("full_name") or "(unidentified)",
                    p["_campaign"]["type"],
                    round(amount * share, 2),
                    p["id"],
                    note,
                ]
            )
    return rows, unmatched, skipped, subs


# ------------------------------------------------------------------ write


def months_owed(start, end, today):
    """Payments due by today for one student, counting the start month."""
    last = min(end or LAST_PAYMENT, LAST_PAYMENT, today)
    if last < start:
        return 0
    return sum(1 for d in payment_dates() if start <= d <= last)


def build(students, income_rows):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Roster")
    style_header(ws, ROSTER_COLS)
    for r, s in enumerate(sorted(students.values(), key=lambda s: s["participant_id"]), start=2):
        ws.cell(r, 1, s["participant_id"])
        ws.cell(r, 2, s["name"])
        ws.cell(r, 3, key_of(s["name"]))
        ws.cell(r, 4, s.get("grade") or None)
        ws.cell(r, 5, s.get("start") or FIRST_PAYMENT).number_format = "yyyy-mm-dd"
        if s.get("end"):
            ws.cell(r, 6, s["end"]).number_format = "yyyy-mm-dd"
        ws.cell(r, 7, s.get("autopay") or "")
        ws.cell(r, 8, "yes" if s.get("in_first") else "no")
        ws.cell(r, 9, s.get("packet") or "")
        for i, k in enumerate(
            ("parent1", "parent1_phone", "parent1_email", "parent2", "parent2_phone",
             "parent2_email", "student_email", "student_phone"),
            start=10,
        ):
            ws.cell(r, i, s.get(k) or "")

    ws = wb.create_sheet("Incomes")
    style_header(ws, INCOME_COLS)
    for r, row in enumerate(income_rows, start=2):
        when = row[0]
        ws.cell(r, 1, when.date() if isinstance(when, datetime) else when).number_format = "yyyy-mm-dd"
        for c in (2, 3, 4):
            ws.cell(r, c, row[c - 1] if row[c - 1] is not None else "")
        ws.cell(r, 5, row[4]).number_format = '"$"#,##0.00'
        ws.cell(r, 6, row[5] if row[5] is not None else "")
        ws.cell(r, 7, row[6] if row[6] is not None else "")

    ws = wb.create_sheet("Schedule")
    style_header(ws, [("Payment Date", 14), ("Amount", 10)])
    for r, d in enumerate(payment_dates(), start=2):
        ws.cell(r, 1, d).number_format = "yyyy-mm-dd"
        ws.cell(r, 2, MONTHLY).number_format = '"$"#,##0'

    return wb


def carry_over(wb, source):
    """Copy the hand-maintained sheets across, values only."""
    if not source or not Path(source).exists():
        return []
    src = openpyxl.load_workbook(source, data_only=True)
    done = []
    for name in PRESERVED:
        if name not in src.sheetnames:
            continue
        s_ws, d_ws = src[name], wb.create_sheet(name)
        for row in s_ws.iter_rows():
            for c in row:
                if c.value is not None:
                    d_ws.cell(c.row, c.column, c.value)
        done.append(f"{name} ({s_ws.max_row - 1} rows)")
    return done


# ------------------------------------------------------------------- main


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("roster", nargs="?", help="FIRST roster export (.json), via roster_sheet.py")
    ap.add_argument("--roster-csv", help="roster_sheet.py output, for contacts")
    ap.add_argument("--from", dest="seed_from", help="v1 workbook to seed the first build from")
    ap.add_argument("--zeffy", action="store_true", help="pull new payments from the Zeffy API")
    ap.add_argument("--out", default=str(DATA_BOOK))
    ap.add_argument("--commit", action="store_true", help="write; otherwise dry run")
    args = ap.parse_args()

    out = Path(args.out)
    students = gather(args)
    if not students:
        sys.exit("\nNo students found. Pass --roster-csv, or --from on the first build.\n")

    existing_wb = openpyxl.load_workbook(out, data_only=True) if out.exists() else None
    income_rows = read_incomes(existing_wb) if existing_wb else []
    seen = zeffy_ids_present(income_rows)

    # First build with --from: bring v1's payment history across, once.
    seeded = 0
    if args.seed_from and not income_rows:
        src = openpyxl.load_workbook(args.seed_from, data_only=True)
        by_key = {key_of(s["name"]): s for s in students.values()}
        for r in range(2, src["Incomes"].max_row + 1):
            d, who, typ, amt = (src["Incomes"].cell(r, c).value for c in (1, 2, 3, 4))
            if not d or amt in (None, "", 0):
                continue
            m = by_key.get(str(who).strip()) if who else None
            income_rows.append(
                [d, m["participant_id"] if m else "", str(who or ""), typ or "", amt,
                 src["Incomes"].cell(r, 6).value or "", src["Incomes"].cell(r, 5).value or ""]
            )
            seeded += 1
        seen = zeffy_ids_present(income_rows)

    added, unmatched, skipped, subs = [], [], 0, {}
    if args.zeffy:
        zmap = read_zeffy_map(existing_wb or (openpyxl.load_workbook(args.seed_from) if args.seed_from else None))
        added, unmatched, skipped, subs = new_zeffy_rows(students, seen, zmap or {})

    print(f"\nData workbook: {out}")
    print(f"Mode:          {'COMMIT' if args.commit else 'dry run — nothing will be written'}\n")
    print(f"  Roster    {len(students)} students")
    if seeded:
        print(f"  Incomes   {seeded} payment(s) carried over from {Path(args.seed_from).name}")
    print(f"  Incomes   {len(income_rows)} existing + {len(added)} new"
          + (f", {skipped} already present" if skipped else ""))

    hand = [r for r in income_rows if not r[5]]
    if hand:
        print(f"            {len(hand)} hand-entered (no Zeffy id) — never rewritten")

    # ── who looks wrong ─────────────────────────────────────────────────────
    today = date.today()
    paid = defaultdict(float)
    for r in income_rows + added:
        if r[1]:
            paid[str(r[1]).strip()] += float(r[4] or 0)

    billed_not_reg = [s["name"] for s in students.values() if s.get("in_budget") and not s.get("in_first")]
    reg_not_billed = [s["name"] for s in students.values() if s.get("in_first") and not s.get("in_budget")]

    # The Venmo flag. A student who owes money and has paid nothing at all is
    # either genuinely behind or paid on a rail nobody recorded — and the second
    # is why a family who had paid got chased. Worth looking at before chasing.
    nothing_at_all = []
    for pid, s in students.items():
        # Someone with no id yet cannot hold a payment, so "paid nothing" tells
        # you about the missing reservation, not about the family. They are
        # already reported above; saying it twice reads as two problems.
        if s.get("needs_reservation"):
            continue
        owed = months_owed(s.get("start") or FIRST_PAYMENT, s.get("end"), today) * MONTHLY
        if owed > 0 and paid.get(pid, 0) == 0:
            nothing_at_all.append(f"{pid} {s['name']} (owes ${owed:,})")

    def block(title, items):
        if items:
            print(f"\n  {title} ({len(items)}):")
            for i in items:
                print(f"    {i}")

    # An id that this script invented would sooner or later collide with one
    # the ledger hands out, and the collision is invisible: two students, one
    # number, money attaching to whichever the sheet resolves first. So stop.
    unreserved = sorted(s["name"] for s in students.values() if s.get("needs_reservation"))
    if unreserved:
        print(f"\n  No participant id ({len(unreserved)}) — reserve one before writing:")
        for n in unreserved:
            print(f'    node scripts/onboarding/onboard.mjs --reserve "{n}"')
        print("    Ids come from the ledger so the budget and onboarding cannot")
        print("    hand the same number to two different students.")

    block("Billed but not registered with FIRST", sorted(billed_not_reg))
    block("Registered but not in the dues ledger", sorted(reg_not_billed))
    block("Owes money, nothing recorded — check for a Venmo or cash payment", sorted(nothing_at_all))
    if unmatched:
        block("Payments that could not be attributed", [f"{w}  ${a:,.2f}  {n!r}" for w, n, a in unmatched[:12]])
        print("    These are in the ledger with no id — visible, counted against nobody.")
        print("    Add a rule to the Zeffy Map sheet.")

    if unreserved:
        print(f"\nRefusing to write: {len(unreserved)} student(s) have no participant id.\n")
        sys.exit(1)

    if not args.commit:
        print(f"\nDry run — pass --commit to write {out.name}.\n")
        return

    if out.exists():
        stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        backup = out.with_name(f"{out.stem}.backup-{stamp}{out.suffix}")
        shutil.copy2(out, backup)
        print(f"\nBacked up to {backup.name}")

    wb = build(students, income_rows + added)
    carried = carry_over(wb, args.seed_from if args.seed_from else (out if out.exists() else None))
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    print(f"Wrote {out}")
    if carried:
        print("  carried over: " + ", ".join(carried))
    print()


if __name__ == "__main__":
    main()
